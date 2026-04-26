"""
Excel Exporter — Network Intelligence Spreadsheet Generator
=============================================================
Exports all network devices and their metadata from the Radar database
into a well-formatted, colour-coded Excel (.xlsx) file for offline
analysis, auditing, or sharing.

Usage:
    make export
    # File saved to: ~/radar_network_audit.xlsx

Standalone:
    PYTHONPATH=. .venv/bin/python3 -m radar.reports.excel_exporter
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def _bytes_to_human(n: int) -> str:
    """Convert raw bytes to a human-readable string (e.g., 1.4 MB)."""
    for unit in ["B", "KB", "MB", "GB"]:
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


class ExcelExporter:
    """
    Queries the Vault database and exports all network device intelligence
    to a formatted Excel spreadsheet.
    """

    def __init__(self, vault=None):
        from radar.database.vault import Vault
        self.vault = vault or Vault()

    def export_devices(self, output_path: Optional[str] = None) -> str:
        """
        Generates the Excel report and saves it to output_path.

        Returns:
            Absolute path to the saved .xlsx file.
        """
        try:
            import pandas as pd
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("Missing dependencies. Run: pip install pandas openpyxl")
            raise

        if output_path is None:
            output_path = str(Path.home() / f"radar_network_audit_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

        logger.info("Querying Vault for network device data...")
        devices = self.vault.get_network_devices()

        if not devices:
            logger.warning("No devices found in database.")

        # ── Build rows ────────────────────────────────────────────────────────
        rows = []
        for d in devices:
            # Calculate how long device has been seen on network
            try:
                duration = d.last_seen - d.first_seen
                days = duration.days
                hours = duration.seconds // 3600
                seen_for = f"{days}d {hours}h" if days > 0 else f"{hours}h"
            except Exception:
                seen_for = "N/A"

            rows.append({
                "IP Address":       d.ip_address,
                "MAC Address":      d.mac_address,
                "Device Name":      d.device_name or "Unknown",
                "Device Type":      d.device_type or "Unknown",
                "Manufacturer":     d.manufacturer or "Unknown",
                "Confidence (%)":   d.confidence,
                "Last Activity":    d.last_activity or "Idle / Passive",
                "DNS Hostname":     d.mdns_hostname or "",
                "NetBIOS Name":     d.netbios_name or "",
                "Traffic Summary":  d.traffic_summary or "",
                "Total Bandwidth":  _bytes_to_human(d.total_bytes or 0),
                "First Seen":       d.first_seen.strftime("%Y-%m-%d %H:%M"),
                "Last Seen":        d.last_seen.strftime("%Y-%m-%d %H:%M"),
                "On Network For":   seen_for,
            })

        df = pd.DataFrame(rows)

        # ── Write to Excel with formatting ────────────────────────────────────
        logger.info(f"Writing {len(rows)} device(s) to {output_path}...")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Network Devices", startrow=2)
            ws = writer.sheets["Network Devices"]

            # ── Title row ─────────────────────────────────────────────────────
            ws["A1"] = f"Radar Network Intelligence Report — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill("solid", fgColor="0D1B2A")
            ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # ── Header row styling ────────────────────────────────────────────
            header_fill = PatternFill("solid", fgColor="1B4F72")
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=3, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[3].height = 22

            # ── Data row styling (alternating colors) ─────────────────────────
            fill_even = PatternFill("solid", fgColor="EBF5FB")
            fill_odd  = PatternFill("solid", fgColor="FFFFFF")
            thin_border = Border(
                left=Side(style="thin", color="D5D8DC"),
                right=Side(style="thin", color="D5D8DC"),
                top=Side(style="thin", color="D5D8DC"),
                bottom=Side(style="thin", color="D5D8DC"),
            )
            for row_idx in range(4, 4 + len(rows)):
                fill = fill_even if row_idx % 2 == 0 else fill_odd
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(vertical="center")

            # ── Auto column widths ────────────────────────────────────────────
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                values = df[col_name].astype(str).tolist() + [col_name]
                max_len = min(max(len(str(v)) for v in values) + 3, 40)
                ws.column_dimensions[col_letter].width = max_len

            # ── Freeze top rows ───────────────────────────────────────────────
            ws.freeze_panes = "A4"

            # ── Auto-filter on header ─────────────────────────────────────────
            ws.auto_filter.ref = ws.dimensions

        logger.info(f"✅ Excel report saved: {output_path}")
        return output_path


# ── Standalone CLI ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    output = sys.argv[1] if len(sys.argv) > 1 else None
    exporter = ExcelExporter()
    path = exporter.export_devices(output_path=output)
    print(f"\n📊 Export complete → {path}")
